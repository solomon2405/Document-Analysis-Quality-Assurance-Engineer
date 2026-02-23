from __future__ import annotations

import asyncio
import logging
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import pdfplumber
from docx import Document
from fastapi import UploadFile
from openpyxl import load_workbook

from app.config import settings
from app.services.ocr_service import run_ocr

logger = logging.getLogger(__name__)

TOKEN_PATTERN = re.compile(r"\S+")
ALLOWED_SUFFIXES = {".docx", ".pdf", ".txt", ".xlsx", ".png", ".jpg", ".jpeg"}


@dataclass(slots=True)
class Token:
    text: str
    file_name: str
    page: int
    paragraph: int
    line: int

    @property
    def location(self) -> str:
        return (
            f"file={self.file_name}, page={self.page}, "
            f"paragraph={self.paragraph}, line={self.line}"
        )


@dataclass(slots=True)
class CombinedText:
    tokens: list[Token]
    full_text: str
    ocr_confidences: list[float]


def _tokenize(text: str, file_name: str, page: int, paragraph: int, line: int) -> list[Token]:
    return [
        Token(
            text=match.group(0),
            file_name=file_name,
            page=page,
            paragraph=paragraph,
            line=line,
        )
        for match in TOKEN_PATTERN.finditer(text)
    ]


def _extract_docx(file_name: str, file_bytes: bytes) -> list[Token]:
    document = Document(BytesIO(file_bytes))
    tokens: list[Token] = []
    for idx, paragraph in enumerate(document.paragraphs, start=1):
        text = (paragraph.text or "").strip()
        if not text:
            continue
        tokens.extend(_tokenize(text=text, file_name=file_name, page=1, paragraph=idx, line=idx))
    return tokens


def _extract_pdf(file_name: str, file_bytes: bytes) -> list[Token]:
    tokens: list[Token] = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line_no, line in enumerate(text.splitlines(), start=1):
                clean = line.strip()
                if clean:
                    tokens.extend(
                        _tokenize(
                            text=clean,
                            file_name=file_name,
                            page=page_no,
                            paragraph=line_no,
                            line=line_no,
                        )
                    )
    return tokens


def _extract_txt(file_name: str, file_bytes: bytes) -> list[Token]:
    decoded = file_bytes.decode("utf-8", errors="ignore")
    tokens: list[Token] = []
    for line_no, line in enumerate(decoded.splitlines(), start=1):
        clean = line.strip()
        if clean:
            tokens.extend(
                _tokenize(
                    text=clean,
                    file_name=file_name,
                    page=1,
                    paragraph=line_no,
                    line=line_no,
                )
            )
    return tokens


def _extract_xlsx(file_name: str, file_bytes: bytes) -> list[Token]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    tokens: list[Token] = []
    paragraph_no = 0

    try:
        for sheet in workbook.worksheets:
            for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if not values:
                    continue
                paragraph_no += 1
                line_text = " ".join(values)
                tokens.extend(
                    _tokenize(
                        text=line_text,
                        file_name=file_name,
                        page=1,
                        paragraph=paragraph_no,
                        line=row_no,
                    )
                )
    finally:
        workbook.close()

    return tokens


def _extract_image(file_name: str, file_bytes: bytes) -> tuple[list[Token], float]:
    ocr_result = run_ocr(file_bytes)
    tokens = [
        Token(
            text=word.text,
            file_name=file_name,
            page=1,
            paragraph=word.line_no,
            line=word.line_no,
        )
        for word in ocr_result.words
    ]
    return tokens, ocr_result.avg_confidence


def _extract_from_file(file_name: str, file_bytes: bytes) -> tuple[list[Token], list[float]]:
    suffix = Path(file_name).suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        raise ValueError(f"Unsupported file type: {suffix}")

    try:
        if suffix == ".docx":
            return _extract_docx(file_name, file_bytes), []
        if suffix == ".pdf":
            return _extract_pdf(file_name, file_bytes), []
        if suffix == ".txt":
            return _extract_txt(file_name, file_bytes), []
        if suffix == ".xlsx":
            return _extract_xlsx(file_name, file_bytes), []

        image_tokens, conf = _extract_image(file_name, file_bytes)
        return image_tokens, [conf]
    except Exception as exc:
        logger.exception("Failed extracting text from %s", file_name)
        raise ValueError(f"Failed to process file '{file_name}': {exc}") from exc


async def combine_files(files: list[UploadFile]) -> CombinedText:
    sorted_files = sorted(files, key=lambda file: (file.filename or "").lower())

    async def process_file(file: UploadFile) -> tuple[list[Token], list[float]]:
        file_bytes = await file.read()
        max_size = settings.max_file_size_mb * 1024 * 1024
        if len(file_bytes) > max_size:
            raise ValueError(
                f"File {file.filename} exceeds {settings.max_file_size_mb}MB limit."
            )
        tokens, confidences = await asyncio.to_thread(
            _extract_from_file, file.filename or "unknown", file_bytes
        )
        logger.info("Processed %s with %s tokens", file.filename, len(tokens))
        return tokens, confidences

    results = await asyncio.gather(*(process_file(file) for file in sorted_files))
    token_collection: list[Token] = []
    ocr_confidences: list[float] = []
    for tokens, confidences in results:
        token_collection.extend(tokens)
        ocr_confidences.extend(confidences)

    full_text = " ".join(token.text for token in token_collection)
    return CombinedText(tokens=token_collection, full_text=full_text, ocr_confidences=ocr_confidences)
